#encoding=utf-8
import xlrd
from xlwt import *
from openpyxl import Workbook as wb
import os
import re


year_list = ['Year1','Year2','Year3','Year4','Year5','Year6','Year7']
year_file_list = ['1114_1165','1166_1217','1218_1269','1270_1321','1322_1373','1374_1426','1427_1478']
fileName0 = 'G:\\数据\\match_table3.xls'
bk=xlrd.open_workbook(fileName0)
shxrange=range(bk.nsheets)
try:
    sh=bk.sheet_by_name("Sheet1")
except:
    print ("代码出错")
ncols=sh.ncols #获取列数
nrows=sh.nrows #获取列数

book = Workbook(encoding='utf-8')
# sheet = book.add_sheet('Sheet1') #创建一个sheet
UPC = []

tmp_add = ['General Purpose Cleaner','Mayonnaise','PeanutButter','Soups']
tmp1_add = ['General Purpose Cleaner','Mayonnaise','PeanutButter','Soups']
tmp2_add = ['hhclean','mayo','peanbutr','soup']
tmp3_add = ['General Purpose Cleaner','Mayonnaise','PeanutButter','soup']



table_head = ['IRI_KEY','WEEK','SY','GE','VEND','ITEM','UNITS','DOLLARS','F','D','PR']

for j in range(0,7):
	for i in range(0,len(tmp1_add)):
		fileName1='G:\\数据\\PLA-UPC文档\\PLA_'+tmp1_add[i]+'_UPC.xls' #UPC码文档
		bk=xlrd.open_workbook(fileName1)
		shxrange=range(bk.nsheets)
		try:
			print('打开成功')
			sh=bk.sheet_by_name("Sheet1")
		except:
			print ("代码出错")


		book = Workbook(encoding='utf-8')
		# sheet = book.add_sheet('Sheet1') #创建一个sheet
		UPC = []
		count = 0
		#print(sh.col_values(18)[5])
		tmp = sh.col_values(1)[1:]
		print(tmp)

		# 得到一个excel中所有的UPC
		for upc_item in tmp:
			if isinstance(upc_item, float):  #若只有一个浮点数且不是短位
				upc_item = re.sub('[^0-9]','',str(int(upc_item)))
				UPC.append(int(upc_item))
			else:  #若是一个列表
				upc_item = upc_item.split(';')
				first_item = upc_item[0]
				high_pos = first_item[0:5]
				for item in upc_item:
					item = re.sub('[^0-9]','',item)
					if item != '':
						if len(item) == 5:
							UPC.append(str(int(high_pos)*100000+int(item)).zfill(5))
						else:
							UPC.append(str(int(item)).zfill(5))
		for ele in UPC:  #去掉所有位数不为10的UPC码
			if len(str(ele))>10 or len(str(ele))<=5:
				UPC.remove(ele)
		for ele in UPC:  #去掉所有位数不为10的UPC码
			if len(str(ele))>10 or len(str(ele))<=5:
				UPC.remove(ele)
		for ele in UPC:  #去掉所有位数不为10的UPC码
			if len(str(ele))>10 or len(str(ele))<=5:
				UPC.remove(ele)
		
		upc_length = len(UPC)





#---------------------------开始匹配杂货店----------------------------

		for test_upc in UPC:
					count += 1
					print('------------'+str(count/upc_length*100)+'%'+'------------')
					print(year_list[j]+'\\'+tmp2_add[i]+year_list[j])
					#print(test_upc)
					book_groc = wb()
					xlsheet = book_groc.get_sheet_by_name('Sheet')
					headlen = len(table_head)
					for k in range(headlen):
						xlsheet.cell(row=1,column=k+1).value = table_head[k] #写表头
					
					fileName3 = open('G:\\数据\\Academic Dataset External copy\\'+year_list[j]+'\\External\\'+tmp2_add[i]+'\\'+tmp2_add[i]+'_groc_'+year_file_list[j])  #销售记录文档
					for line in fileName3:
						line=line.strip('\n')
						line=line.split()
						if "IRI_KEY" in line:
							continue
						if len(line[4]) < 5:
							line[4] = line[4] + (5-len(line[4]))*'0'
						UPC_record = int(line[4]) * 100000 + int(line[5])
						if str(test_upc) in str(UPC_record):
							print(line)
							xlsheet.append(line)
					if (xlsheet.cell(row=2,column=1).value != None):
						isExists = os.path.exists('G:\\数据\\result\\'+year_list[j]+'\\'+tmp2_add[i]+'\\groc')
						if not isExists:
							os.makedirs('G:\\数据\\result\\'+year_list[j]+'\\'+tmp2_add[i]+'\\groc')
							book_groc.save('G:\\数据\\result\\'+year_list[j]+'\\'+tmp2_add[i]+'\\groc\\'+year_list[j]+'-'+tmp2_add[i]+'groc'+'-'+str(test_upc)+'.xlsx')
						else:
							book_groc.save('G:\\数据\\result\\'+year_list[j]+'\\'+tmp2_add[i]+'\\groc\\'+year_list[j]+'-'+tmp2_add[i]+'groc'+'-'+str(test_upc)+'.xlsx')
		fileName3.close();