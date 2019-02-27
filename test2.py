#encoding=utf-8
import xlrd
from xlwt import *
from openpyxl import Workbook as wb
#------------------读数据---------------------------------
fileName1='G:\\数据\\code\\PLA_Beer_UPC_20010101-20071230.xls' #UPC码文档
bk=xlrd.open_workbook(fileName1)
shxrange=range(bk.nsheets)
try:
    sh=bk.sheet_by_name("View Results1")
except:
    print ("代码出错")
ncols=sh.ncols #获取列数
nrows=sh.nrows #获取列数

book = Workbook(encoding='utf-8')
# sheet = book.add_sheet('Sheet1') #创建一个sheet
UPC = []
tmp = sh.col_values(17)[5:]

# 得到一个excel中所有的UPC
for upc_item in tmp:
	if isinstance(upc_item, float):  #若只有一个浮点数
		UPC.append(int(upc_item))
	else:  #若是一个列表
		upc_item = upc_item.split(';')
		for item in upc_item:
			UPC.append(int(item))



type_total = 0;
for test_upc in UPC:
	# print('------------------------------------------------------')
	table_head = ['IRI_KEY','WEEK','SY','GE','VEND','ITEM','UNITS','DOLLARS','F','D','PR']
	book2 = wb()
	xlsheet = book2.get_sheet_by_name('Sheet')
	headlen = len(table_head)
	for i in range(headlen):
		xlsheet.cell(row=1,column=i+1).value = table_head[i] #写表头

	fileName2 = open('G:\\数据\\code\\beer_drug_1114_1165')  #销售记录文档
	for line in fileName2:
		line=line.strip('\n')
		line=line.split()
		if "IRI_KEY" in line:
			continue
		if len(line[4]) < 5:
			line[4] = line[4] + (5-len(line[4]))*'0'
		UPC_record = int(line[4]) * 100000 + int(line[5])
		if str(test_upc) in str(UPC_record):
			print(line)
			xlsheet.append(line)

	if (xlsheet.cell(row=2,column=1).value != None):
		book2.save("result"+str(test_upc)+'.xlsx')

fileName2.close();
